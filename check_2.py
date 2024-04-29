import pandas as pd
from tkinter import *
from tkinter import ttk  # Import ttk for Treeview widget
from openpyxl import load_workbook
#-----------------------------------------------------------#
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch

from reportlab.platypus import Image as ReportLabImage
#-------------------------------------------------------------#

from reportlab.platypus import SimpleDocTemplate, Table, TableStyle,  Paragraph, Spacer

from reportlab.lib.styles import getSampleStyleSheet
import datetime

def main_fun():
    def export_to_pdf(data):
        timestamp = datetime.datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
        pdf_file = f'{timestamp}.pdf'
        doc = SimpleDocTemplate(pdf_file, pagesize=A4, rightMargin=18, leftMargin=18, topMargin=18)
        elements = []

        extra_image_path = 'test.png'  # Replace with the actual path to your image

        extra_img = ReportLabImage(extra_image_path, width=50, height=50)
        elements.append(extra_img)

        # text_content for generating pdf
        text_content = (
            f"Hello, In this PDF for power factor value calculation by the user by the date and time of {timestamp}\n"
            f"Primary Value: {primary.get()}\n"
            f"Secondary Value: {secondary.get()}\n"
            f"Average Value: {result_var.get()}")



        text_paragraph = Paragraph(text_content, getSampleStyleSheet()['BodyText'])
        elements.append(text_paragraph)

        spacer = Spacer(1, 12)  # Adjust the height of the spacer as needed (1 inch = 72 points)
        elements.append(spacer)

        # Define table data and style
        table_data = [['S.no', 'Power', 'OP', 'CT/R', 'Above / Below', 'KVA', 'kW', 'KVAr']]
        # Add more rows to table_data if needed
        table_data.extend(data)  # Add data rows

        table_style = TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                                  ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                                  ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                                  ('GRID', (0, 0), (-1, -1), 1, (0, 0, 0))])

        # Create table
        table = Table(table_data)
        table.setStyle(table_style)
        elements.append(table)

        # Build PDF
        doc.build(elements)
        done_lab = Label(root, text="PDF Generated Successfully!!!").place(x=350, y=700)

    def export_to_excel(data):
        df = pd.DataFrame(data, columns=['S.no', 'Power', 'OP', 'CT/R', 'Above / Below', 'KVA', 'kW', 'KVAr'])
        output_file = 'output.xlsx'
        df.to_excel(output_file, index=False)

    def update_result(*args):
        try:
            value1 = float(primary.get())
            value2 = float(secondary.get())
            result_var.set(value1 / value2)

            # Load the Excel file into a DataFrame
            excel_file = 'data.xlsx'
            df = pd.read_excel(excel_file)

            # Define the input value
            input_value = float(value1 / value2)  # Use the calculated result as the input value

            # Compare 'OP' column with the input value
            df['Comparison'] = df['OP'].apply(
                lambda x: 'Condition-2' if x > input_value else ('Condition-1' if x < input_value else 'Equal'))

            df['KVA'] = df['OP'].apply(
                lambda x: 'Condition-2' if x > input_value else ('Condition-1' if x < input_value else 'Equal'))

            df['kW'] = df['OP'].apply(
                lambda x: 'Condition-2' if x > input_value else ('Condition-1' if x < input_value else 'Equal'))

            df['KVAr'] = df['OP'].apply(
                lambda x: 'Condition-2' if x > input_value else ('Condition-1' if x < input_value else 'Equal'))

            # Save the DataFrame back to Excel with the results in a new column
            output_file = 'output.xlsx'  # Change this to the desired output file path
            df.to_excel(output_file, index=False)

            # Load the updated Excel file
            wb2 = load_workbook(output_file)
            ws = wb2.active

            # Get data from specific columns
            data = []
            for row in ws.iter_rows(min_row=1, max_col=8, values_only=True):
                data.append(row)

            # Clear existing treeview data
            for item in treeview.get_children():
                treeview.delete(item)

            # Insert data into treeview
            for item in data[1:]:
                treeview.insert("", "end", values=item)
            return value1
            return value2
            return result_var

            #print("Data loaded successfully.")
        except ValueError:
            result_var.set("Enter both two values!!!")

    def export_data():
        data = []
        for item in treeview.get_children():
            values = treeview.item(item, 'values')
            data.append(values)

        export_to_pdf(data)
        export_to_excel(data)

    root = Tk()
    root.title("Project")
    root.geometry("800x800")
    root.resizable(False, False)

    primary = StringVar()
    secondary = StringVar()
    result_var = StringVar()

    prim_label = Label(root, text="Enter the Primary value")
    prim_label.place(x=350, y=20)
    Secon_label = Label(root, text="Enter the Secondary value")
    Secon_label.place(x=350, y=80)
    prim_entry = Entry(root, textvariable=primary)
    prim_entry.place(x=350, y=50)
    secon_entry = Entry(root, textvariable=secondary)
    secon_entry.place(x=350, y=110)
    result_label = Label(root, textvariable=result_var)
    result_label.place(x=360, y=130)

    # Create a Frame for Treeview and Scrollbar
    frame = Frame(root)
    frame.place(x=10, y=200, width=760, height=500)

    # Create a Treeview widget for displaying data in a table
    treeview = ttk.Treeview(frame, columns=('S.no', 'Power', 'OP', 'CT/R', 'Condition', 'KVA', 'kW', 'KVAr'),
                            show='headings')
    treeview.pack(side=LEFT, fill=BOTH, expand=True)


    # Create a Scrollbar
    scrollbar = Scrollbar(root, orient=VERTICAL, command=treeview.yview)
    scrollbar.place(x=775, y=200, height = 500)

    scrollbar_1 = Scrollbar(root, orient=HORIZONTAL, command=treeview.xview)
    scrollbar_1.place(x=0, y= 700, width = 800)

    # Configure the Treeview to use the Scrollbars
    treeview.config(yscrollcommand=scrollbar.set, xscrollcommand=scrollbar_1.set)

    # Define headings for the columns
    treeview.heading('S.no', text='S.no', anchor='center')
    treeview.heading('Power', text='Power', anchor='center')
    treeview.heading('OP', text='OP', anchor='center')
    treeview.heading('CT/R', text='CT/R', anchor='center')
    treeview.heading('Condition', text='Condition', anchor='center')
    treeview.heading('KVA', text='KVA', anchor='center')
    treeview.heading('kW', text='kW', anchor='center')
    treeview.heading('KVAr', text='KVAr', anchor='center')


    # Set column widths and anchor to center
    treeview.column('S.no', width=50, anchor='center')
    treeview.column('Power', width=70, anchor='center')
    treeview.column('OP', width=70, anchor='center')
    treeview.column('CT/R', width=70, anchor='center')
    treeview.column('Condition', width=100, anchor='center')
    treeview.column('KVA', width=70, anchor='center')
    treeview.column('kW', width=70, anchor='center')
    treeview.column('KVAr', width=70, anchor='center')


    # Bind the update_result function to changes in primary and secondary values
    primary.trace_add('write', update_result)
    secondary.trace_add('write', update_result)
    #
    print_btn = Button(root, text = "Print", command=export_data).place(x=375, y=750)

    root.mainloop()

main_fun()