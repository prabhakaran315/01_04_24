'''import pandas as pd
from tkinter import *
from tkinter import ttk  # Import ttk for Treeview widget
from openpyxl import load_workbook

def main_fun():
    def update_result(*args):
        try:
            value1 = float(primary.get())
            value2 = float(secondary.get())
            result_var.set(value1 / value2)

            # Load the Excel file into a DataFrame
            excel_file = 'data.xlsx'

            df = pd.read_excel(excel_file)

            # Define the input value
            input_value = float(value1 / value2)  # Use the calculated result as the input value

            # Compare 'OP' column with the input value
            df['Comparison'] = df['OP'].apply(
                lambda x: 'Below' if x > input_value else ('Above' if x < input_value else 'both'))

            df['KVA'] = df['OP'].apply(
                lambda x: 'Below' if x > input_value else ('Above' if x < input_value else 'both'))

            df['kW'] = df['OP'].apply(
                lambda x: 'Below' if x > input_value else ('Above' if x < input_value else 'both'))

            df['KVAr'] = df['OP'].apply(
                lambda x: 'Below' if x > input_value else ('Above' if x < input_value else 'both'))

            # Save the DataFrame back to Excel with the results in a new column
            output_file = 'output.xlsx'  # Change this to the desired output file path
            df.to_excel(output_file, index=False)

            # Load the updated Excel file
            wb2 = load_workbook(output_file)
            ws = wb2.active

            # Get data from specific columns
            data = []
            for row in ws.iter_rows(min_row=1, max_col=8, values_only=True):
                data.append(row)

            # Clear existing treeview data
            for item in treeview.get_children():
                treeview.delete(item)

            # Insert data into treeview
            for item in data[1:]:
                treeview.insert("", "end", values=item)

            #print("Data loaded successfully.")
        except ValueError:
            result_var.set("Enter both two values!!!")

    root = Tk()
    root.title("Project")
    root.geometry("800x800")
    root.resizable(False, False)

    primary = StringVar()
    secondary = StringVar()
    result_var = StringVar()

    prim_label = Label(root, text="Enter the Primary value")
    prim_label.place(x=350, y=20)
    Secon_label = Label(root, text="Enter the Secondary value")
    Secon_label.place(x=350, y=80)
    prim_entry = Entry(root, textvariable=primary)
    prim_entry.place(x=350, y=50)
    secon_entry = Entry(root, textvariable=secondary)
    secon_entry.place(x=350, y=110)
    result_label = Label(root, textvariable=result_var)
    result_label.place(x=360, y=130)

    # Create a Frame for Treeview and Scrollbar
    frame = Frame(root)
    frame.place(x=10, y=200, width=780, height=500)

    # Create a Treeview widget for displaying data in a table
    treeview = ttk.Treeview(frame, columns=('S.no', 'Power', 'OP', 'CT/R', 'Above / Below', 'KVA', 'kW', 'KVAr'),
                            show='headings')
    treeview.pack(side=LEFT, fill=BOTH, expand=True)

    # Create a Scrollbar
    scrollbar = Scrollbar(frame, orient=VERTICAL, command=treeview.yview)
    scrollbar.pack(side=RIGHT, fill=Y)

    scrollbar_1 = Scrollbar(frame, orient=HORIZONTAL, command=treeview.xview)
    scrollbar_1.pack(side=BOTTOM, fill=X)

    # Configure the Treeview to use the Scrollbar
    treeview.config(yscrollcommand=scrollbar.set)
    treeview.config(xscrollcommand=scrollbar_1.set)
    # Define headings for the columns
    treeview.heading('S.no', text='S.no')
    treeview.heading('Power', text='Power')
    treeview.heading('OP', text='OP')
    treeview.heading('CT/R', text='CT/R')
    treeview.heading('Above / Below', text='Above / Below')
    treeview.heading('KVA', text='KVA')
    treeview.heading('kW', text='kW')
    treeview.heading('KVAr', text='KVAr')

    # Set column widths
    treeview.column('S.no', width=50, anchor=CENTER)
    treeview.column('Power', width=70, anchor=CENTER)
    treeview.column('OP', width=70, anchor=CENTER)
    treeview.column('CT/R', width=70, anchor=CENTER)
    treeview.column('Above / Below', width=100, anchor=CENTER)
    treeview.column('KVA', width=70, anchor=CENTER)
    treeview.column('kW', width=70, anchor=CENTER)
    treeview.column('KVAr', width=70, anchor=CENTER)

    # Bind the update_result function to changes in primary and secondary values
    primary.trace_add('write', update_result)
    secondary.trace_add('write', update_result)

    print_btn = Button(root, text = "Print").place(x=375, y=750)

    root.mainloop()

main_fun()
'''
import tkinter as tk
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import Table, Spacer, Paragraph, BaseDocTemplate, Frame, PageTemplate, TableStyle
from reportlab.lib import colors
from tkinter import messagebox
import math
from reportlab.lib.units import inch
import datetime

root = tk.Tk()
root.title("Project")

frame_1 = tk.Frame(root, bg="Grey")
frame_1.pack(expand=True, fill=tk.BOTH)
frame_2 = tk.Frame(root, height=650, bg="White")
frame_2.pack(expand=True, fill=tk.BOTH)
frame_3 = tk.Frame(root, height=80, bg="Grey")
frame_3.pack(expand=True, fill=tk.BOTH)

primary = tk.StringVar()
secondary = tk.StringVar()

def update_result(*args):
    for widget in frame_2.winfo_children():
        widget.destroy()
    for widget_1 in frame_3.winfo_children():
        widget_1.destroy()

def submit():
    try:
        value_1 = float(primary_entry.get())
        value_2 = float(secondary_entry.get())

        if value_1 > 0 and value_2 > 0:
            cal_opt_val = round(value_1 / value_2)
            condition_1(value_1, value_2, cal_opt_val)
        else:
            messagebox.showinfo('Error!', "Kindly enter positive integers and greater than zero values!!!")
    except ValueError:
        messagebox.showinfo("Error!", "Kindly enter both values as numbers!!!")

def condition_1(value_1, value_2, cal_opt_val):
    global hello
    global first_three
    global data_2
    kva = math.floor(1.732 * 415.0 * (value_1 / 1000))
    kvar = math.ceil(kva * 0.03)
    kw = math.ceil(kva * 0.05)
    data_1 = [
            ["ID", "Panel Rating", "Optimal CT Ratio", "Condition", "Minimum kVAr", "Minimum kW"],
            [1, "630A ASTRA - B 3", 1252, "", "", ""],
            [2, "630A ASTRA - B 5", 1809, "", "", ""],
            [3, "630A ASTRA - B10", 2782, "", "", ""],
            [4, "315A ASTRA - B 3", 835, "", "", ""],
            [5, "315A ASTRA - B 5", 1252, "", "", ""],
            [6, "315A ASTRA - B10", 1530, "", "", ""],
            [7, "420A ASTRA - B10", 1669, "", "", ""],
            [8, "420A ASTRA - B 5", 1252, "", "", ""],
            [9, "210A ASTRA - B10", 1530, "", "", ""],
            [10, "210A ASTRA - B 5", 696, "", "", ""],
            [11, "150A ASTRA - B 5", 835, "", "", ""],
            [12, "150A ASTRA - B10", 1669, "", "", ""]
        ]
    data_3 = [
            ["Optimal kW"],
            [225],
            [325],
            [500],
            [150],
            [225],
            [275],
            [300],
            [225],
            [275],
            [125],
            [150],
            [300]
        ]

    data_2 = [row[:] for row in data_1]
    for row_idx, row in enumerate(data_2[1:], start=1):
        data_val = row[2]
        if data_val < cal_opt_val:
            row[3:5] = "Condition - 1", kvar, kw
        else:
            if row_idx < len(data_3):
                row[3:5] = "Condition - 2", "-", data_3[row_idx][0]

    for widget in frame_2.winfo_children():
        widget.destroy()

    for x in range(len(data_2)):
        for y in range(len(data_2[0])):
            font_style = ("Arial", 15, "bold") if x == 0 else ("Arial", 15)
            tk.Label(frame_2, text=data_2[x][y], width=20, anchor="center", borderwidth=1, relief="solid", font=font_style).grid(row=x, column=y, padx=0, pady=0, sticky="nsew")

    first_three = []
    sort_column = sorted(data_2, key=lambda x: str(x[5]))
    first = sort_column[:3]
    for row in sort_column:
        if str(row[5]) in [str(item[5]) for item in first]:
            first_three.append((row[0], row[1], row[5]))

    table_frame = tk.Frame(frame_2, bg="white")
    table_frame.grid(row=len(data_1) + 1, column=0, columnspan=3, pady=10)

    tk.Label(table_frame, text="Optimal Panel Rating Based on Optimum kW", font=("Arial", 16, "bold"), bg="White").grid(row=0, column=0, columnspan=5, padx=9, pady=5)
    tk.Label(table_frame, text="Panel ID ", font=("Arial", 14, "bold"), bg="White").grid(row=1, column=0, padx=5, pady=5)
    tk.Label(table_frame, text="Panel Rating", font=("Arial", 14, "bold"), bg="White").grid(row=1, column=1, padx=5, pady=5)
    tk.Label(table_frame, text="Minimum kW", font=("Arial", 14, "bold"), bg="White").grid(row=1, column=2, padx=5, pady=5)

    row_index = 2
    for i in range(min(3, len(first_three))):
        panel_id = first_three[i][0]
        panel_rating = first_three[i][1]
        panel_kw = first_three[i][2]
        pan_rat = (panel_id, panel_rating, panel_kw)

        category, data, data_1 = pan_rat

        tk.Label(table_frame, text=category, font=("Arial", 14), bg="White").grid(row=row_index, column=0, padx=5, pady=5)
        tk.Label(table_frame, text=str(data), font=("Arial", 14), bg="White").grid(row=row_index, column=1, padx=5, pady=5)
        tk.Label(table_frame, text=str(data_1), font=("Arial", 14), bg="White").grid(row=row_index, column=2, padx=5, pady=5)
        row_index += 1

    hello = tk.IntVar()
    tk.Checkbutton(frame_3, text="", variable=hello, onvalue=1, offvalue=0, bg="Grey").place(relx=0.42, rely=0.22, anchor=tk.NW)
    tk.Label(frame_3, text=' If you want to print suggestion panel rating!!!', bg="Grey", fg="White", font=("Arial", 14)).place(relx=0.43, rely=0.2, anchor=tk.NW)
    tk.Button(frame_3, text="Print", command=export_pdf_condition_1, font=("Arial", 15, "bold")).place(relx=0.5, rely=0.5, anchor=tk.NW)

def header(canvas, doc):
    canvas.saveState()
    canvas.drawImage('test.png', 40, 770, width=100, height=50)
    canvas.restoreState()

    now = timestamp = datetime.datetime.now().strftime("%d/%m/%Y")
    header_style = ParagraphStyle(name='HeaderStyle', fontSize=12, textColor='black')
    header_text = Paragraph(f"{now}", header_style)
    header_text.wrapOn(canvas, inch, inch)
    header_text.drawOn(canvas, 500, 800)

def export_pdf_condition_1():
    global first_three
    global hello
    global data_2
    timestamp = datetime.datetime.now().strftime("%d_%m_%Y_%H_%M_%S")
    pdf_filename = f'{timestamp}.pdf'

    doc = BaseDocTemplate(pdf_filename, pagesize=A4)
    header_frame = Frame(doc.leftMargin, doc.topMargin, doc.width, doc.height)
    header_template = PageTemplate(id='header_template', frames=[header_frame], onPage=header)
    doc.addPageTemplates([header_template])

    styles = getSampleStyleSheet()
    title_style = styles['Title']
    body_style = styles['BodyText']

    now = datetime.datetime.now().strftime("%d/%m/%Y")
    header_text = Paragraph(f"Date: {now}", body_style)

    elements = [
        header_text,
        Spacer(1, 12),
        Paragraph("True Power Factor", title_style),
        Spacer(1, 12),
        Paragraph("The true power factor is a measure of how efficiently electrical power is being used. "
                "It represents the ratio of true power (measured in watts) to apparent power (measured in volt-amperes). "
                "A higher true power factor indicates better utilization of electrical power.", body_style),
        Spacer(1, 20)
    ]

    primary_value_paragraph = Paragraph('Primary value : {}'.format(primary_entry.get()), body_style)
    secondary_value_paragraph = Paragraph("Secondary value : {}".format(secondary_entry.get()), body_style)
    elements.append(primary_value_paragraph)
    elements.append(secondary_value_paragraph)
    elements.append(Spacer(1, 20))

    data = [["ID", "Panel Rating", "Optimal CT Ratio", "Condition", "Minimum kVAr", "Minimum kW"]]
    data.extend([item[:4] + [item[4], item[5]] for item in data_2[1:]])

    t = Table(data, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, -1), (-1, 0), 20),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)]
    ))

    elements += [t]

    check_box = hello.get()
    if check_box == 1:
        try:
            content_1 = [
                Paragraph("Optimal Panel Rating Based on Optimum kW", title_style), Spacer(1, 5)
            ]
            elements.extend(content_1)

            table_data = [["Panel ID", "Panel Rating", "Minimum kW"]]
            for i in range(3):
                panel_id = first_three[i][0]
                panel_rating = first_three[i][1]
                panel_kw = first_three[i][2]
                pan_rat = (panel_id, panel_rating, panel_kw)

                for row in [pan_rat]:
                    table_data.append(list(row))

            table = Table(table_data, repeatRows=1)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, -1), (-1, 0), 20),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)]
            ))
            elements.append(table)

        except Exception as e:
            print("Error:", e)

        doc.build(elements)
        messagebox.showinfo("Done", "PDF file exported Successfully!!!")

    else:
        doc.build(elements)
        messagebox.showinfo("Done", "PDF file exported Successfully!!!")

# Labels
tk.Label(frame_1, text="Conditions to Achieve True Power Factor Performance", bg="Grey", fg="White", font=("Arial", 20, "bold")).pack(padx=10, pady=10)
tk.Label(frame_1, text="Enter the Primary value of transformer", bg="Grey", fg="White", font=("Arial", 14)).pack()
primary_entry = tk.Entry(frame_1, textvariable=primary, font=("Arial", 15))
primary_entry.pack()

tk.Label(frame_1, text="Enter the Secondary value of transformer", bg="Grey", fg="White", font=("Arial", 14)).pack()
secondary_entry = tk.Entry(frame_1, textvariable=secondary, font=("Arial", 15))
secondary_entry.pack()

primary.trace_add('write', update_result)
secondary.trace_add('write', update_result)

# Buttons
tk.Button(frame_1, text="Submit", command=submit, font=("Arial", 15, "bold")).pack(padx=10, pady=10)

root.mainloop()